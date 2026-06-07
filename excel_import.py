#!/usr/bin/env python3
"""Import bar inventory data from an Excel file.

Usage:
  python excel_import.py inventory.xlsx       # import data
  python excel_import.py --template           # generate a blank template

The Excel file needs up to 5 sheets (all optional, any order):
  Suppliers, Products, Stock, Catalog, Budget

Column names are matched flexibly — German and English variants work,
capitalisation and extra spaces are ignored. See --template for examples.
"""

import sys
import re
import httpx
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUR_BASE = "http://localhost:8000"

# ── Flexible column name matching ─────────────────────────────────────────────
# Each field maps to a list of accepted column header variants (lowercase).
FIELD_ALIASES: dict[str, list[str]] = {
    # Supplier fields
    "name":            ["name", "supplier name", "lieferant", "bezeichnung", "firma"],
    "contact_email":   ["email", "e-mail", "mail", "contact email", "kontakt", "e_mail"],
    "wallet_address":  ["wallet", "wallet address", "algorand", "algo address", "address", "adresse"],
    # Product fields
    "unit":            ["unit", "einheit", "unit type", "einheitentyp", "typ"],
    "preferred_supplier": ["preferred supplier", "lieferant", "bevorzugter lieferant",
                           "supplier", "default supplier"],
    # Stock fields
    "product":         ["product", "produkt", "artikel", "item", "product name", "produktname"],
    "quantity":        ["quantity", "qty", "menge", "bestand", "lagerbestand",
                        "current stock", "aktueller bestand", "anzahl"],
    "reorder_point":   ["reorder point", "reorder_point", "mindestbestand",
                        "nachbestellpunkt", "meldebestand", "minimum stock"],
    "reorder_qty":     ["reorder qty", "reorder_qty", "reorder quantity", "nachbestellmenge",
                        "bestellmenge", "order quantity"],
    "max_price":       ["max price", "max_price", "maximum price", "höchstpreis",
                        "max preis", "preis limit", "price limit"],
    # Catalog fields
    "supplier":        ["supplier", "lieferant", "supplier name", "lieferantenname"],
    "unit_price":      ["unit price", "unit_price", "preis", "price", "stückpreis",
                        "einzelpreis", "price per unit", "preis pro einheit"],
    "min_order_qty":   ["min order qty", "min_order_qty", "min order", "mindestmenge",
                        "minimum order", "mindestbestellmenge"],
}

# Sheet name matching
SHEET_ALIASES = {
    "suppliers": ["suppliers", "lieferanten", "supplier", "lieferant", "vendors"],
    "products":  ["products", "produkte", "product", "produkt", "items", "artikel"],
    "stock":     ["stock", "lager", "lagerbestand", "inventory", "bestand"],
    "catalog":   ["catalog", "catalogue", "katalog", "preisliste", "price list",
                  "supplier catalog", "lieferantenkatalog"],
    "budget":    ["budget", "budget", "etat", "gesamtbudget"],
}


def _normalise(s: str) -> str:
    return re.sub(r"[\s_\-]+", " ", s.strip().lower())


def _match_field(header: str, *field_names: str) -> str | None:
    h = _normalise(header)
    for field in field_names:
        for alias in FIELD_ALIASES.get(field, [field]):
            if h == _normalise(alias):
                return field
    return None


def _find_sheet(wb: Workbook, sheet_type: str) -> openpyxl.worksheet.worksheet.Worksheet | None:
    aliases = SHEET_ALIASES.get(sheet_type, [sheet_type])
    for name in wb.sheetnames:
        if _normalise(name) in [_normalise(a) for a in aliases]:
            return wb[name]
    return None


def _sheet_to_dicts(ws) -> list[dict]:
    """Convert a worksheet to a list of dicts using the first row as headers."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # skip blank rows
        result.append({headers[i]: row[i] for i in range(len(headers))})
    return result


def _map_columns(rows: list[dict], *fields: str) -> list[dict]:
    """Remap arbitrary column names to our canonical field names."""
    if not rows:
        return []
    sample_headers = list(rows[0].keys())
    mapping = {}  # original header → canonical field name
    for h in sample_headers:
        field = _match_field(h, *fields)
        if field:
            mapping[h] = field
    return [{mapping[k]: v for k, v in row.items() if k in mapping} for row in rows]


# ── API helpers ───────────────────────────────────────────────────────────────
def _post(path: str, body: dict) -> dict | None:
    try:
        r = httpx.post(f"{OUR_BASE}{path}", json=body, timeout=10)
        r.raise_for_status()
        return r.json()
    except httpx.ConnectError:
        print(f"\n  ERROR: backend not reachable at {OUR_BASE}")
        print("  Start it with:  uvicorn backend.main:app --reload\n")
        sys.exit(1)
    except httpx.HTTPStatusError as e:
        print(f"    ✗ HTTP {e.response.status_code}: {e.response.text[:120]}")
        return None


def _put(path: str, body: dict) -> dict | None:
    r = httpx.put(f"{OUR_BASE}{path}", json=body, timeout=10)
    r.raise_for_status()
    return r.json()


def _get(path: str) -> list | dict:
    r = httpx.get(f"{OUR_BASE}{path}", timeout=10)
    r.raise_for_status()
    return r.json()


def _random_algo_address() -> str:
    try:
        from algosdk.account import generate_account
        _, address = generate_account()
        return address
    except ImportError:
        return "XNFT2AMETQIRMODE3BKK4PPYH3HH4OKFPMFIKMSZMOAPIYUJLQJYIFEA4"


def _val(row: dict, field: str, default=None):
    v = row.get(field, default)
    if v is None or str(v).strip() == "":
        return default
    return v


# ── Import logic ──────────────────────────────────────────────────────────────
def import_excel(path: str):
    print(f"🍸  bar-inventory ← Excel import  ({path})\n")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"  ERROR: file not found: {path}\n")
        sys.exit(1)

    print(f"  Sheets found: {', '.join(wb.sheetnames)}\n")

    # Verify backend
    try:
        _get("/health")
    except Exception:
        print("  ERROR: backend not reachable — run: uvicorn backend.main:app --reload\n")
        sys.exit(1)

    existing_suppliers = {s["name"]: s for s in _get("/inventory/suppliers")}
    existing_products  = {p["name"]: p for p in _get("/inventory/products")}
    existing_stock     = {s["product_id"]: s for s in _get("/inventory/stock")}
    existing_catalog   = {(c["supplier_id"], c["product_id"]) for c in _get("/inventory/catalog")}

    created_suppliers: dict[str, dict] = {}
    created_products:  dict[str, dict] = {}

    # ── Budget ────────────────────────────────────────────────────────────────
    ws_budget = _find_sheet(wb, "budget")
    if ws_budget:
        print("  [Budget]")
        rows = list(ws_budget.iter_rows(values_only=True))
        budget_val = None
        for row in rows:
            for cell in row:
                try:
                    v = float(cell)
                    if v > 0:
                        budget_val = v
                        break
                except (TypeError, ValueError):
                    pass
            if budget_val:
                break
        if budget_val:
            _put("/inventory/budget", {"total_budget": budget_val})
            print(f"    ✓ Budget set to {budget_val:.2f} ALGO\n")
        else:
            print("    ✗ No numeric budget value found\n")
    else:
        print("  [Budget] sheet not found — skipping\n")

    # ── Suppliers ─────────────────────────────────────────────────────────────
    ws_sup = _find_sheet(wb, "suppliers")
    if ws_sup:
        print("  [Suppliers]")
        rows = _map_columns(
            _sheet_to_dicts(ws_sup),
            "name", "contact_email", "wallet_address",
        )
        for row in rows:
            name = _val(row, "name")
            if not name:
                continue
            name = str(name).strip()
            if name in existing_suppliers:
                print(f"    → {name} (already exists)")
                created_suppliers[name] = existing_suppliers[name]
                continue
            sup = _post("/inventory/suppliers", {
                "name":           name,
                "contact_email":  str(_val(row, "contact_email", f"orders@{name.lower().replace(' ','')}.example")),
                "wallet_address": str(_val(row, "wallet_address") or _random_algo_address()),
            })
            if sup:
                created_suppliers[name] = sup
                existing_suppliers[name] = sup
                print(f"    ✓ {name}")
        print()
    else:
        print("  [Suppliers] sheet not found — skipping\n")
        created_suppliers = existing_suppliers

    # ── Products ──────────────────────────────────────────────────────────────
    ws_prod = _find_sheet(wb, "products")
    if ws_prod:
        print("  [Products]")
        rows = _map_columns(
            _sheet_to_dicts(ws_prod),
            "name", "unit", "preferred_supplier",
        )
        for row in rows:
            name = _val(row, "name")
            if not name:
                continue
            name = str(name).strip()
            if name in existing_products:
                print(f"    → {name} (already exists)")
                created_products[name] = existing_products[name]
                continue
            pref_sup_name = _val(row, "preferred_supplier")
            pref_sup_id   = None
            if pref_sup_name:
                match = (existing_suppliers | created_suppliers).get(str(pref_sup_name).strip())
                pref_sup_id = match["id"] if match else None

            prod = _post("/inventory/products", {
                "name":                  name,
                "unit":                  str(_val(row, "unit", "unit")).strip(),
                "preferred_supplier_id": pref_sup_id,
            })
            if prod:
                created_products[name] = prod
                existing_products[name] = prod
                print(f"    ✓ {name}  [{_val(row, 'unit', 'unit')}]")
        print()
    else:
        print("  [Products] sheet not found — skipping\n")
        created_products = existing_products

    all_products = existing_products | created_products

    # ── Stock ─────────────────────────────────────────────────────────────────
    ws_stock = _find_sheet(wb, "stock")
    if ws_stock:
        print("  [Stock]")
        rows = _map_columns(
            _sheet_to_dicts(ws_stock),
            "product", "quantity", "reorder_point", "reorder_qty", "max_price",
        )
        for row in rows:
            prod_name = _val(row, "product")
            if not prod_name:
                continue
            prod_name = str(prod_name).strip()
            product   = all_products.get(prod_name)
            if not product:
                print(f"    ✗ '{prod_name}' — product not found (import Products sheet first)")
                continue
            pid = product["id"]
            if pid in existing_stock:
                print(f"    → {prod_name} (stock already exists)")
                continue
            qty   = int(_val(row, "quantity",     0) or 0)
            rp    = int(_val(row, "reorder_point", 5) or 5)
            rq    = int(_val(row, "reorder_qty",  12) or 12)
            mp    = float(_val(row, "max_price",  0.05) or 0.05)
            stock = _post("/inventory/stock", {
                "product_id":    pid,
                "quantity":      qty,
                "reorder_point": rp,
                "reorder_qty":   rq,
                "max_price":     mp,
            })
            if stock:
                existing_stock[pid] = stock
                status = "LOW ⚠" if qty <= rp else "OK ✓"
                print(f"    ✓ {prod_name}: qty={qty} reorder@{rp}  [{status}]")
        print()
    else:
        print("  [Stock] sheet not found — skipping\n")

    # ── Catalog ───────────────────────────────────────────────────────────────
    ws_cat = _find_sheet(wb, "catalog")
    if ws_cat:
        print("  [Catalog]")
        rows = _map_columns(
            _sheet_to_dicts(ws_cat),
            "supplier", "product", "unit_price", "min_order_qty",
        )
        for row in rows:
            sup_name  = _val(row, "supplier")
            prod_name = _val(row, "product")
            if not sup_name or not prod_name:
                continue
            sup_name  = str(sup_name).strip()
            prod_name = str(prod_name).strip()

            supplier = (existing_suppliers | created_suppliers).get(sup_name)
            product  = all_products.get(prod_name)
            if not supplier:
                print(f"    ✗ Supplier '{sup_name}' not found")
                continue
            if not product:
                print(f"    ✗ Product '{prod_name}' not found")
                continue

            sid, pid = supplier["id"], product["id"]
            if (sid, pid) in existing_catalog:
                print(f"    → {sup_name} × {prod_name} (already exists)")
                continue

            price = float(_val(row, "unit_price", 0.05) or 0.05)
            moq   = int(_val(row, "min_order_qty", 1) or 1)
            result = _post("/inventory/catalog", {
                "supplier_id":   sid,
                "product_id":    pid,
                "unit_price":    price,
                "min_order_qty": moq,
            })
            if result:
                existing_catalog.add((sid, pid))
                print(f"    ✓ {sup_name} × {prod_name}  @ {price:.4f} ALGO/unit")
        print()
    else:
        print("  [Catalog] sheet not found — skipping\n")

    print("  Import complete. Open http://localhost:5173 to see the data.\n")


# ── Template generator ────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="4F46E5")   # indigo
EXAMPLE_FILL  = PatternFill("solid", fgColor="EEF2FF")   # light indigo
HEADER_FONT   = Font(color="FFFFFF", bold=True)
EXAMPLE_FONT  = Font(color="6B7280", italic=True)
BORDER_SIDE   = Side(style="thin", color="D1D5DB")
CELL_BORDER   = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)


def _header_row(ws, headers: list[str]):
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.font       = HEADER_FONT
        cell.fill       = HEADER_FILL
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = CELL_BORDER
    ws.row_dimensions[1].height = 20


def _example_row(ws, row_num: int, values: list):
    for col, v in enumerate(values, 1):
        cell           = ws.cell(row=row_num, column=col, value=v)
        cell.font      = EXAMPLE_FONT
        cell.fill      = EXAMPLE_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border    = CELL_BORDER


def _auto_width(ws, extra: int = 4):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + extra


def generate_template(path: str = "bar_inventory_template.xlsx"):
    wb = Workbook()

    # ── README ────────────────────────────────────────────────────────────────
    ws_readme = wb.active
    ws_readme.title = "README"
    notes = [
        ("Bar Inventory — Excel Import Template", True),
        ("", False),
        ("Fill in the other sheets and run:", False),
        ("  python excel_import.py bar_inventory_template.xlsx", False),
        ("", False),
        ("Rules:", True),
        ("• Grey rows are examples — replace or delete them", False),
        ("• Column names are flexible: German / English / mixed", False),
        ("• Sheets you don't need can be left blank or deleted", False),
        ("• Supplier / Product names must match exactly across sheets", False),
        ("• Wallet Address: leave blank to auto-generate an Algorand address", False),
        ("• Budget: any positive number in the Budget sheet works", False),
        ("• Unit Price and Max Price are in ALGO", False),
    ]
    for i, (text, bold) in enumerate(notes, 1):
        cell      = ws_readme.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=11 if bold else 10)
    ws_readme.column_dimensions["A"].width = 65

    # ── Suppliers ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Suppliers")
    _header_row(ws, ["Name", "Email", "Wallet Address"])
    _example_row(ws, 2, ["The Gin House", "orders@ginhouse.example", ""])
    _example_row(ws, 3, ["Alpine Brewery", "orders@alpine.example", ""])
    _auto_width(ws)

    # ── Products ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Products")
    _header_row(ws, ["Name", "Unit", "Preferred Supplier"])
    _example_row(ws, 2, ["Hendrick's Gin", "bottle", "The Gin House"])
    _example_row(ws, 3, ["Aperol",         "bottle", "The Gin House"])
    _example_row(ws, 4, ["Pilsner",        "case",   "Alpine Brewery"])
    _example_row(ws, 5, ["Tonic Water",    "case",   "Alpine Brewery"])
    _auto_width(ws)

    # ── Stock ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Stock")
    _header_row(ws, ["Product", "Quantity", "Reorder Point", "Reorder Qty", "Max Price (ALGO)"])
    _example_row(ws, 2, ["Hendrick's Gin", 2,  5,  12, 0.05])
    _example_row(ws, 3, ["Aperol",         18, 6,  12, 0.04])
    _example_row(ws, 4, ["Pilsner",        3,  10, 24, 0.03])
    _example_row(ws, 5, ["Tonic Water",    20, 8,  24, 0.02])
    _auto_width(ws)

    # ── Catalog ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Catalog")
    _header_row(ws, ["Supplier", "Product", "Unit Price (ALGO)", "Min Order Qty"])
    _example_row(ws, 2, ["The Gin House",  "Hendrick's Gin", 0.048, 6])
    _example_row(ws, 3, ["The Gin House",  "Aperol",         0.035, 6])
    _example_row(ws, 4, ["Alpine Brewery", "Pilsner",        0.022, 12])
    _example_row(ws, 5, ["Alpine Brewery", "Tonic Water",    0.014, 12])
    _auto_width(ws)

    # ── Budget ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Budget")
    ws.cell(row=1, column=1, value="Total Budget (ALGO)").font = Font(bold=True)
    cell       = ws.cell(row=2, column=1, value=100.0)
    cell.font  = EXAMPLE_FONT
    cell.fill  = EXAMPLE_FILL
    ws.column_dimensions["A"].width = 22

    wb.save(path)
    print(f"\n  Template saved to: {path}")
    print("  Fill in the sheets and run:")
    print(f"    python excel_import.py {path}\n")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2 or sys.argv[1] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    if sys.argv[1] == "--template":
        out = sys.argv[2] if len(sys.argv) > 2 else "bar_inventory_template.xlsx"
        generate_template(out)
    else:
        import_excel(sys.argv[1])
